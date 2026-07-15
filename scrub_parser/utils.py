"""
Utility functions for the NCD Scrub Dashboard parser.
Filename date extraction, worksheet helpers, and date math.
"""

from __future__ import annotations

import datetime
import logging
import re
from decimal import Decimal, InvalidOperation
from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Filename → date
# ---------------------------------------------------------------------------

DATE_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")


def extract_date_from_filename(filename: str) -> Optional[datetime.date]:
    """
    Parse MM.DD.YYYY from a filename string.
    Returns None if no valid date is found.
    """
    m = DATE_RE.search(filename)
    if not m:
        return None
    month, day, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
    try:
        return datetime.date(year, month, day)
    except ValueError:
        logger.warning("Invalid date components in filename '%s': %s/%s/%s",
                       filename, month, day, year)
        return None


# ---------------------------------------------------------------------------
# Worksheet navigation helpers
# ---------------------------------------------------------------------------

def find_sheet(wb, *candidates: str) -> Optional[Worksheet]:
    """
    Return the first worksheet whose name matches any of the candidates
    (case-insensitive). Falls back to the first visible sheet if none match.
    """
    lower_map = {s.title.strip().lower(): s for s in wb.worksheets}
    for name in candidates:
        ws = lower_map.get(name.strip().lower())
        if ws is not None:
            return ws
    # fallback: first visible sheet
    for s in wb.worksheets:
        if s.sheet_state == "visible":
            return s
    return wb.worksheets[0] if wb.worksheets else None


def find_summary_sheet(wb) -> Optional[Worksheet]:
    """Convenience: find the Summary sheet."""
    return find_sheet(wb, "Summary", "summary", "SUMMARY")


def find_data_sheet(wb) -> Optional[Worksheet]:
    """Convenience: find the primary data sheet (not Summary)."""
    for ws in wb.worksheets:
        title = ws.title.strip().lower()
        if title not in ("summary", "instructions", "notes", "config"):
            return ws
    return wb.worksheets[0] if wb.worksheets else None


def cell_text(cell) -> str:
    """Return stripped string representation of a cell value, or ''."""
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def cell_number(cell) -> Optional[int | float]:
    """Return numeric value from a cell, or None."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    # try parsing string
    try:
        cleaned = str(v).replace(",", "").replace("$", "").strip()
        if "." in cleaned:
            return float(cleaned)
        return int(cleaned)
    except (ValueError, TypeError):
        return None


def cell_decimal(cell) -> Optional[Decimal]:
    """Return Decimal value from a cell, or None."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    try:
        cleaned = str(v).replace(",", "").replace("$", "").strip()
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None


# ---------------------------------------------------------------------------
# Row scanning
# ---------------------------------------------------------------------------

def find_row_by_label(ws: Worksheet,
                      label: str,
                      col: int = 1,
                      start_row: int = 1,
                      end_row: Optional[int] = None,
                      exact: bool = False) -> Optional[int]:
    """
    Scan a column for a row whose cell text contains `label` (case-insensitive).
    Returns the row number, or None.

    If exact=True, requires the cell text to match the label exactly
    (still case-insensitive).
    """
    max_row = end_row or ws.max_row or 1
    label_lower = label.lower()
    for row in range(start_row, max_row + 1):
        txt = cell_text(ws.cell(row=row, column=col)).lower()
        if exact and txt == label_lower:
            return row
        if not exact and label_lower in txt:
            return row
    return None


def find_all_rows_by_label(ws: Worksheet,
                           label: str,
                           col: int = 1,
                           start_row: int = 1,
                           end_row: Optional[int] = None) -> list[int]:
    """Return ALL row numbers where column `col` contains `label`."""
    max_row = end_row or ws.max_row or 1
    label_lower = label.lower()
    rows = []
    for row in range(start_row, max_row + 1):
        txt = cell_text(ws.cell(row=row, column=col)).lower()
        if label_lower in txt:
            rows.append(row)
    return rows


def count_data_rows(ws: Worksheet, header_row: int = 1) -> int:
    """
    Count non-empty rows below the header row.
    A row is 'non-empty' if at least one cell in columns A-E has a value.
    """
    count = 0
    max_row = ws.max_row or 1
    for row in range(header_row + 1, max_row + 1):
        has_data = any(
            ws.cell(row=row, column=c).value is not None
            for c in range(1, min(6, (ws.max_column or 5) + 1))
        )
        if has_data:
            count += 1
    return count


def sum_column(ws: Worksheet,
               col: int,
               start_row: int,
               end_row: Optional[int] = None) -> int:
    """Sum numeric values in a column range."""
    max_row = end_row or ws.max_row or 1
    total = 0
    for row in range(start_row, max_row + 1):
        n = cell_number(ws.cell(row=row, column=col))
        if n is not None and isinstance(n, (int, float)):
            total += int(n)
    return total


# ---------------------------------------------------------------------------
# Section boundary detection
# ---------------------------------------------------------------------------

def find_section_boundaries(ws: Worksheet,
                            section_labels: list[str],
                            col: int = 1) -> dict[str, tuple[int, int]]:
    """
    Given a list of section header labels, find the start and end rows
    for each section. Returns {label: (start_row, end_row)}.

    Sections are bounded by the next section header or the last row.
    """
    max_row = ws.max_row or 1
    found: list[tuple[str, int]] = []

    for row in range(1, max_row + 1):
        txt = cell_text(ws.cell(row=row, column=col)).lower()
        for label in section_labels:
            if label.lower() in txt:
                found.append((label, row))
                break

    boundaries: dict[str, tuple[int, int]] = {}
    for i, (label, start) in enumerate(found):
        if i + 1 < len(found):
            end = found[i + 1][1] - 1
        else:
            end = max_row
        boundaries[label] = (start, end)

    return boundaries


# ---------------------------------------------------------------------------
# Date math
# ---------------------------------------------------------------------------

def months_between(d1: datetime.date, d2: datetime.date) -> float:
    """Approximate months between two dates (d1 - d2)."""
    return (d1 - d2).days / 30.44
