"""
Parser functions for each NCD scrub type.

Every parser takes an openpyxl Workbook (and optionally a file_date for
date-math scrubs) and returns dict[str, int | Decimal] mapping output keys
to extracted values.

A parser raises ParserError if the file structure is unexpected.
"""

from __future__ import annotations

import datetime
import logging
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook

from scrub_parser.utils import (
    cell_decimal,
    cell_number,
    cell_text,
    count_data_rows,
    find_all_rows_by_label,
    find_data_sheet,
    find_row_by_label,
    find_section_boundaries,
    find_summary_sheet,
    months_between,
    sum_column,
)

logger = logging.getLogger(__name__)


class ParserError(Exception):
    """Raised when a file can't be parsed (missing sheet, unexpected layout)."""


# ── helpers ────────────────────────────────────────────────────────────────

def _require_summary(wb: Workbook, scrub_name: str):
    ws = find_summary_sheet(wb)
    if ws is None:
        raise ParserError(f"{scrub_name}: no Summary sheet found")
    return ws


def _grand_total_value(ws, col_offset: int = 1,
                       start_row: int = 1,
                       end_row: Optional[int] = None,
                       label: str = "Grand Total") -> Optional[int]:
    """Find 'Grand Total' in column A, return the value from col A + offset."""
    row = find_row_by_label(ws, label, col=1, start_row=start_row, end_row=end_row)
    if row is None:
        return None
    return cell_number(ws.cell(row=row, column=1 + col_offset))


def _total_value(ws, col_offset: int = 1,
                 start_row: int = 1,
                 end_row: Optional[int] = None) -> Optional[int]:
    """Find 'Total' (exact) in column A, return adjacent value."""
    row = find_row_by_label(ws, "Total", col=1, start_row=start_row,
                            end_row=end_row, exact=True)
    if row is None:
        # try non-exact as fallback
        row = find_row_by_label(ws, "Total", col=1, start_row=start_row,
                                end_row=end_row)
    if row is None:
        return None
    return cell_number(ws.cell(row=row, column=1 + col_offset))


def _safe_int(v) -> int:
    if v is None:
        return 0
    return int(v)


# ── 1. Billing Date Alignment ─────────────────────────────────────────────

def parse_billing_date_alignment(wb: Workbook,
                                 file_date: Optional[datetime.date] = None
                                 ) -> dict[str, int | Decimal]:
    """
    Summary sheet → Grand Total in column A → column B value.
    Output key: 'count'
    """
    ws = _require_summary(wb, "Billing Date Alignment")
    val = _grand_total_value(ws, col_offset=1)
    if val is None:
        raise ParserError("Billing Date Alignment: Grand Total not found on Summary sheet")
    return {"count": _safe_int(val)}


# ── 2. FBD in the Past + Blank FBD ────────────────────────────────────────

def parse_fbd(wb: Workbook,
              file_date: Optional[datetime.date] = None
              ) -> dict[str, int | Decimal]:
    """
    Summary sheet has two sections:
      - 'FBD in the Past' → Grand Total → count
      - 'FBD Blank'       → Grand Total → count
    Output keys: 'fbd_past', 'fbd_blank'
    """
    ws = _require_summary(wb, "FBD")
    sections = find_section_boundaries(
        ws, ["FBD in the Past", "FBD Blank", "Blank FBD"]
    )

    results: dict[str, int | Decimal] = {}

    # FBD in the Past
    if "FBD in the Past" in sections:
        start, end = sections["FBD in the Past"]
        val = _grand_total_value(ws, start_row=start, end_row=end)
        results["fbd_past"] = _safe_int(val)
    else:
        logger.warning("FBD: 'FBD in the Past' section not found — trying full-sheet Grand Total")
        gt_rows = find_all_rows_by_label(ws, "Grand Total")
        if gt_rows:
            results["fbd_past"] = _safe_int(cell_number(ws.cell(row=gt_rows[0], column=2)))

    # FBD Blank / Blank FBD
    blank_key = "FBD Blank" if "FBD Blank" in sections else "Blank FBD"
    if blank_key in sections:
        start, end = sections[blank_key]
        val = _grand_total_value(ws, start_row=start, end_row=end)
        results["fbd_blank"] = _safe_int(val)
    else:
        # fallback: second Grand Total on the sheet
        gt_rows = find_all_rows_by_label(ws, "Grand Total")
        if len(gt_rows) >= 2:
            results["fbd_blank"] = _safe_int(cell_number(ws.cell(row=gt_rows[1], column=2)))

    return results


# ── 3. NBD in the Past + Blank NBD ────────────────────────────────────────

def parse_nbd(wb: Workbook,
              file_date: Optional[datetime.date] = None
              ) -> dict[str, int | Decimal]:
    """
    Same pattern as FBD — Summary with two sections.
    Output keys: 'nbd_past', 'nbd_blank'
    """
    ws = _require_summary(wb, "NBD")
    sections = find_section_boundaries(
        ws, ["NBD in the Past", "NBD Blank", "Blank NBD"]
    )

    results: dict[str, int | Decimal] = {}

    if "NBD in the Past" in sections:
        start, end = sections["NBD in the Past"]
        val = _grand_total_value(ws, start_row=start, end_row=end)
        results["nbd_past"] = _safe_int(val)
    else:
        gt_rows = find_all_rows_by_label(ws, "Grand Total")
        if gt_rows:
            results["nbd_past"] = _safe_int(cell_number(ws.cell(row=gt_rows[0], column=2)))

    blank_key = "NBD Blank" if "NBD Blank" in sections else "Blank NBD"
    if blank_key in sections:
        start, end = sections[blank_key]
        val = _grand_total_value(ws, start_row=start, end_row=end)
        results["nbd_blank"] = _safe_int(val)
    else:
        gt_rows = find_all_rows_by_label(ws, "Grand Total")
        if len(gt_rows) >= 2:
            results["nbd_blank"] = _safe_int(cell_number(ws.cell(row=gt_rows[1], column=2)))

    return results


# ── 4. Mass Terms ──────────────────────────────────────────────────────────

def parse_mass_terms(wb: Workbook,
                     file_date: Optional[datetime.date] = None
                     ) -> dict[str, int | Decimal]:
    """
    Summary → pivot table → Grand Total row:
      Column B = Count of Member ID
      Column C = Sum of Product Amount
    Output keys: 'count', 'amount'
    """
    ws = _require_summary(wb, "Mass Terms")
    row = find_row_by_label(ws, "Grand Total", col=1)
    if row is None:
        raise ParserError("Mass Terms: Grand Total not found")

    count = cell_number(ws.cell(row=row, column=2))
    amount = cell_decimal(ws.cell(row=row, column=3))

    results: dict[str, int | Decimal] = {"count": _safe_int(count)}
    if amount is not None:
        results["amount"] = amount
    return results


# ── 5. ACH Rebill ─────────────────────────────────────────────────────────

def parse_ach_rebill(wb: Workbook,
                     file_date: Optional[datetime.date] = None
                     ) -> dict[str, int | Decimal]:
    """
    Summary → 'Total Records' row → count.
    Also looks for dollar recapture if present.
    Output keys: 'total_records', 'recapture_amount'
    """
    ws = _require_summary(wb, "ACH Rebill")

    row = find_row_by_label(ws, "Total Records", col=1)
    if row is None:
        raise ParserError("ACH Rebill: 'Total Records' not found on Summary")

    count = cell_number(ws.cell(row=row, column=2))
    results: dict[str, int | Decimal] = {"total_records": _safe_int(count)}

    # look for a dollar recapture field
    recap_row = find_row_by_label(ws, "Recapture", col=1)
    if recap_row is None:
        recap_row = find_row_by_label(ws, "Total Billing", col=1)
    if recap_row is not None:
        amt = cell_decimal(ws.cell(row=recap_row, column=2))
        if amt is not None:
            results["recapture_amount"] = amt

    return results


# ── 6. Incomplete Accounts ────────────────────────────────────────────────

def parse_incomplete_accounts(wb: Workbook,
                              file_date: Optional[datetime.date] = None
                              ) -> dict[str, int | Decimal]:
    """
    Summary → Action/Count table → 'Total' row → column B.
    Output key: 'count'
    """
    ws = _require_summary(wb, "Incomplete Accounts")
    val = _total_value(ws, col_offset=1)
    if val is None:
        val = _grand_total_value(ws, col_offset=1)
    if val is None:
        raise ParserError("Incomplete Accounts: no Total/Grand Total found")
    return {"count": _safe_int(val)}


# ── 7. No Products on Dependents ──────────────────────────────────────────

def parse_no_dependent_products(wb: Workbook,
                                file_date: Optional[datetime.date] = None
                                ) -> dict[str, int | Decimal]:
    """
    Summary → 'Summary of Actions' → Count column → 'Total' or last row.
    Output key: 'count'
    """
    ws = _require_summary(wb, "No Products on Dependents")

    # try standard Total row first
    val = _total_value(ws, col_offset=1)
    if val is not None:
        return {"count": _safe_int(val)}

    # fallback: Grand Total
    val = _grand_total_value(ws, col_offset=1)
    if val is not None:
        return {"count": _safe_int(val)}

    # last resort: sum all numeric values in column B below row 1
    total = sum_column(ws, col=2, start_row=2)
    if total > 0:
        return {"count": total}

    raise ParserError("No Dependent Products: could not determine total count")


# ── 8. Missing MOP ────────────────────────────────────────────────────────

def parse_missing_mop(wb: Workbook,
                      file_date: Optional[datetime.date] = None
                      ) -> dict[str, int | Decimal]:
    """
    Summary → 'Total' row → count.
    Fallback: count data rows on data sheet.
    Output key: 'count'
    """
    ws = find_summary_sheet(wb)
    if ws is not None:
        val = _total_value(ws, col_offset=1)
        if val is None:
            val = _grand_total_value(ws, col_offset=1)
        if val is not None:
            return {"count": _safe_int(val)}

    # fallback: count data rows
    data_ws = find_data_sheet(wb)
    if data_ws is None:
        raise ParserError("Missing MOP: no Summary total and no data sheet found")
    return {"count": count_data_rows(data_ws)}


# ── 9. SRs Aged 30+ Days ─────────────────────────────────────────────────

def parse_srs_aged(wb: Workbook,
                   file_date: Optional[datetime.date] = None
                   ) -> dict[str, int | Decimal]:
    """
    Summary → Action Count table → sum all action counts, or find Total row.
    Output key: 'count'
    """
    ws = _require_summary(wb, "SRs Aged 30+ Days")

    # try Total / Grand Total first
    val = _total_value(ws, col_offset=1)
    if val is not None:
        return {"count": _safe_int(val)}

    val = _grand_total_value(ws, col_offset=1)
    if val is not None:
        return {"count": _safe_int(val)}

    # fallback: sum column B (skip header row)
    total = sum_column(ws, col=2, start_row=2)
    if total > 0:
        return {"count": total}

    raise ParserError("SRs Aged 30+ Days: could not determine total count")


# ── 10. Test Member Accounts ──────────────────────────────────────────────

def parse_test_member_accounts(wb: Workbook,
                               file_date: Optional[datetime.date] = None
                               ) -> dict[str, int | Decimal]:
    """
    Summary → 'Total' row → count.
    Output key: 'count'
    """
    ws = _require_summary(wb, "Test Member Accounts")
    val = _total_value(ws, col_offset=1)
    if val is None:
        val = _grand_total_value(ws, col_offset=1)
    if val is None:
        raise ParserError("Test Member Accounts: no Total/Grand Total found")
    return {"count": _safe_int(val)}


# ── 11. Overaged Dependents ───────────────────────────────────────────────

def parse_overaged_dependents(wb: Workbook,
                              file_date: Optional[datetime.date] = None
                              ) -> dict[str, int | Decimal]:
    """
    Summary → action count table → total.
    May produce split output for Final Notice vs 30-Day Notice if both
    sections exist; otherwise returns a single 'count'.

    Output keys: 'count' (always), 'final_notice' and 'thirty_day' (if split)
    """
    ws = _require_summary(wb, "Overaged Dependents")

    # check for Final Notice / 30-Day sections
    sections = find_section_boundaries(
        ws, ["Final Notice", "30-Day Notice", "30 Day Notice"]
    )

    results: dict[str, int | Decimal] = {}

    if len(sections) >= 2:
        for label, (start, end) in sections.items():
            val = _total_value(ws, start_row=start, end_row=end)
            if val is None:
                val = _grand_total_value(ws, start_row=start, end_row=end)
            key = "final_notice" if "final" in label.lower() else "thirty_day"
            results[key] = _safe_int(val)
        results["count"] = sum(v for v in results.values() if isinstance(v, int))
    else:
        # single section — total or sum
        val = _total_value(ws, col_offset=1)
        if val is None:
            val = _grand_total_value(ws, col_offset=1)
        if val is None:
            val = sum_column(ws, col=2, start_row=2)
        results["count"] = _safe_int(val)

    return results


# ── 12. All Hold Reasons (Policies on Hold 2+ Months) ─────────────────────

def parse_all_hold_reasons(wb: Workbook,
                           file_date: Optional[datetime.date] = None
                           ) -> dict[str, int | Decimal]:
    """
    FILTERED count: count data rows where Hold Date is > 2 months before file_date.
    Output key: 'count'
    """
    if file_date is None:
        raise ParserError("All Hold Reasons: file_date is required for date-math filtering")

    data_ws = find_data_sheet(wb)
    if data_ws is None:
        raise ParserError("All Hold Reasons: no data sheet found")

    # find the Hold Date column
    hold_col = None
    header_row = 1
    for col in range(1, (data_ws.max_column or 20) + 1):
        txt = cell_text(data_ws.cell(row=header_row, column=col)).lower()
        if "hold" in txt and "date" in txt:
            hold_col = col
            break

    if hold_col is None:
        # fallback: count all data rows
        logger.warning("All Hold Reasons: 'Hold Date' column not found — counting all rows")
        return {"count": count_data_rows(data_ws)}

    cutoff = file_date - datetime.timedelta(days=60)
    count = 0
    max_row = data_ws.max_row or 1

    for row in range(header_row + 1, max_row + 1):
        raw = data_ws.cell(row=row, column=hold_col).value
        if raw is None:
            continue
        if isinstance(raw, datetime.datetime):
            hold_date = raw.date()
        elif isinstance(raw, datetime.date):
            hold_date = raw
        else:
            continue
        if hold_date <= cutoff:
            count += 1

    return {"count": count}


# ── 13. $0 Product Fee ────────────────────────────────────────────────────

def parse_zero_product_fee(wb: Workbook,
                           file_date: Optional[datetime.date] = None
                           ) -> dict[str, int | Decimal]:
    """
    Count data rows on the data sheet (no summary tab).
    Note: split files are handled at the runner level (sum across parts).
    Output key: 'count'
    """
    data_ws = find_data_sheet(wb)
    if data_ws is None:
        raise ParserError("$0 Product Fee: no data sheet found")
    return {"count": count_data_rows(data_ws)}


# ── 14. Daily Dupes ───────────────────────────────────────────────────────

def parse_daily_dupes(wb: Workbook,
                      file_date: Optional[datetime.date] = None
                      ) -> dict[str, int | Decimal]:
    """
    Count data rows on the data sheet (no summary tab).
    Output key: 'count'
    """
    data_ws = find_data_sheet(wb)
    if data_ws is None:
        raise ParserError("Daily Dupes: no data sheet found")
    return {"count": count_data_rows(data_ws)}


# ── 15. Combined Eligibility Errors ───────────────────────────────────────

def parse_combined_eligibility(wb: Workbook,
                               file_date: Optional[datetime.date] = None
                               ) -> dict[str, int | Decimal]:
    """
    Summary sheet has sections for MetLife, VSP, NWFA (and optionally OED).
    Sections may be stacked vertically or arranged side-by-side.
    Each section has error types with a 'Split Error Count' column.
    Sum per carrier section.

    Output keys: 'metlife', 'vsp', 'nwfa', 'oed'
    """
    ws = _require_summary(wb, "Combined Eligibility Errors")
    results: dict[str, int | Decimal] = {}

    # ── Strategy 1: section boundaries (vertically stacked sections) ──────
    section_labels = ["MetLife EDI Errors", "VSP EDI Errors", "NWFA Errors", "OED"]
    sections = find_section_boundaries(ws, section_labels)

    if sections:
        for label, (start, end) in sections.items():
            # normalize label → output key
            label_lower = label.lower()
            if "metlife" in label_lower:
                key = "metlife"
            elif "vsp" in label_lower:
                key = "vsp"
            elif "nwfa" in label_lower:
                key = "nwfa"
            elif "oed" in label_lower:
                key = "oed"
            else:
                continue

            gt = _grand_total_value(ws, start_row=start, end_row=end)
            if gt is not None:
                results[key] = _safe_int(gt)
            else:
                # sum the count column within the section
                results[key] = sum_column(ws, col=2, start_row=start + 1, end_row=end)

    if results:
        return results

    # ── Strategy 2: side-by-side column scanning ──────────────────────────
    max_row = ws.max_row or 1
    max_col = ws.max_column or 20

    carrier_configs = [
        ("metlife", ["metlife", "met life"]),
        ("vsp", ["vsp"]),
        ("nwfa", ["nwfa"]),
        ("oed", ["oed", "missing_oed", "missing oed"]),
    ]

    carrier_columns: dict[str, tuple[int, int]] = {}

    for row in range(1, min(6, max_row + 1)):
        for col in range(1, max_col + 1):
            txt = cell_text(ws.cell(row=row, column=col)).lower()
            for key, patterns in carrier_configs:
                if key not in carrier_columns:
                    if any(p in txt for p in patterns):
                        count_col = _find_count_col_near(ws, row, col, max_col)
                        if count_col is not None:
                            carrier_columns[key] = (col, count_col)
                        break

    for key, (header_col, count_col) in carrier_columns.items():
        header_row = _find_header_row(ws, count_col)
        total = sum_column(ws, col=count_col, start_row=header_row + 1)
        results[key] = total

    return results


def _find_count_col_near(ws, row: int, start_col: int, max_col: int) -> Optional[int]:
    """Find column with 'count' or 'split error count' in header near start_col."""
    for col in range(start_col, min(start_col + 10, max_col + 1)):
        for r in range(row, min(row + 4, (ws.max_row or 1) + 1)):
            txt = cell_text(ws.cell(row=r, column=col)).lower()
            if "count" in txt or "split error" in txt:
                return col
    return None


def _find_header_row(ws, col: int) -> int:
    """Find the row containing the header for a given column (first non-empty text row)."""
    for row in range(1, min(10, (ws.max_row or 1) + 1)):
        txt = cell_text(ws.cell(row=row, column=col))
        if txt:
            return row
    return 1


# ── 16. Realm Health Errors ───────────────────────────────────────────────

def parse_realm_health(wb: Workbook,
                       file_date: Optional[datetime.date] = None
                       ) -> dict[str, int | Decimal]:
    """
    Data sheet header indicates carrier (MetLife or VSP).
    Count data rows.
    Output keys: 'metlife_carrier' or 'vsp_carrier'
    """
    data_ws = find_data_sheet(wb)
    if data_ws is None:
        raise ParserError("Realm Health: no data sheet found")

    # determine carrier from headers
    carrier = None
    for row in range(1, min(4, (data_ws.max_row or 1) + 1)):
        for col in range(1, min(10, (data_ws.max_column or 5) + 1)):
            txt = cell_text(data_ws.cell(row=row, column=col)).lower()
            if "metlife" in txt:
                carrier = "metlife_carrier"
                break
            elif "vsp" in txt:
                carrier = "vsp_carrier"
                break
        if carrier:
            break

    if carrier is None:
        logger.warning("Realm Health: could not determine carrier from headers, defaulting to 'metlife_carrier'")
        carrier = "metlife_carrier"

    count = count_data_rows(data_ws)
    return {carrier: count}


# ── 17. Account Updater ──────────────────────────────────────────────────

def parse_account_updater(wb: Workbook,
                          file_date: Optional[datetime.date] = None
                          ) -> dict[str, int | Decimal]:
    """
    Billing Summary → 'Total Members' = count, 'Total Billing' = dollar amount.
    Output keys: 'count', 'amount'
    """
    ws = _require_summary(wb, "Account Updater")

    results: dict[str, int | Decimal] = {}

    # Total Members
    row = find_row_by_label(ws, "Total Members", col=1)
    if row is not None:
        val = cell_number(ws.cell(row=row, column=2))
        results["count"] = _safe_int(val)
    else:
        raise ParserError("Account Updater: 'Total Members' not found")

    # Total Billing
    row = find_row_by_label(ws, "Total Billing", col=1)
    if row is not None:
        amt = cell_decimal(ws.cell(row=row, column=2))
        if amt is not None:
            results["amount"] = amt

    return results


# ── Registry ──────────────────────────────────────────────────────────────

PARSER_REGISTRY: dict[str, callable] = {
    "parse_billing_date_alignment": parse_billing_date_alignment,
    "parse_fbd": parse_fbd,
    "parse_nbd": parse_nbd,
    "parse_mass_terms": parse_mass_terms,
    "parse_ach_rebill": parse_ach_rebill,
    "parse_incomplete_accounts": parse_incomplete_accounts,
    "parse_no_dependent_products": parse_no_dependent_products,
    "parse_missing_mop": parse_missing_mop,
    "parse_srs_aged": parse_srs_aged,
    "parse_test_member_accounts": parse_test_member_accounts,
    "parse_overaged_dependents": parse_overaged_dependents,
    "parse_all_hold_reasons": parse_all_hold_reasons,
    "parse_zero_product_fee": parse_zero_product_fee,
    "parse_daily_dupes": parse_daily_dupes,
    "parse_combined_eligibility": parse_combined_eligibility,
    "parse_realm_health": parse_realm_health,
    "parse_account_updater": parse_account_updater,
}


def get_parser(name: str):
    """Retrieve a parser function by its registered name."""
    func = PARSER_REGISTRY.get(name)
    if func is None:
        raise KeyError(f"No parser registered as '{name}'. "
                       f"Available: {sorted(PARSER_REGISTRY.keys())}")
    return func
