"""
Scrub Dashboard Runner — main orchestrator.

Coordinates:
  1. SharePoint file discovery (via Graph API)
  2. Date extraction from filenames
  3. Dedup check against Postgres
  4. File download + Excel parsing
  5. Result storage

Can be run for all scrub types, a single type, or a date range.
"""

from __future__ import annotations

import datetime
import io
import logging
import re
from collections import defaultdict
from decimal import Decimal
from typing import Optional

from openpyxl import load_workbook

from scrub_parser.config import SCRUB_CONFIGS, ScrubTypeConfig, get_config
from scrub_parser.database import Database
from scrub_parser.models import ParsedFile, ScrubResult, Section
from scrub_parser.parsers import ParserError, get_parser
from scrub_parser.sharepoint import GraphClient, SharePointFile
from scrub_parser.utils import extract_date_from_filename

logger = logging.getLogger(__name__)


class RunStats:
    """Accumulates stats across a run."""

    def __init__(self):
        self.files_found = 0
        self.files_parsed = 0
        self.files_skipped_dedup = 0
        self.files_errored = 0
        self.entries_inserted = 0
        self.entries_skipped = 0
        self.errors: list[str] = []

    def summary(self) -> str:
        lines = [
            f"Files found:        {self.files_found}",
            f"Files parsed:       {self.files_parsed}",
            f"Files skipped (dup): {self.files_skipped_dedup}",
            f"Files errored:      {self.files_errored}",
            f"Entries inserted:   {self.entries_inserted}",
            f"Entries skipped:    {self.entries_skipped}",
        ]
        if self.errors:
            lines.append(f"\nErrors ({len(self.errors)}):")
            for e in self.errors[:20]:
                lines.append(f"  - {e}")
            if len(self.errors) > 20:
                lines.append(f"  ... and {len(self.errors) - 20} more")
        return "\n".join(lines)


class Runner:
    """
    Main orchestrator for the scrub parser pipeline.

    Usage:
        runner = Runner(graph_client, database)
        stats = runner.run_all()
        print(stats.summary())
    """

    def __init__(self, graph: GraphClient, db: Database, *,
                 dry_run: bool = False):
        self.graph = graph
        self.db = db
        self.dry_run = dry_run

    # ── Public API ─────────────────────────────────────────────────────────

    def run_all(self, *,
                start_date: Optional[datetime.date] = None,
                end_date: Optional[datetime.date] = None) -> RunStats:
        """Run all configured scrub types."""
        stats = RunStats()
        for cfg in SCRUB_CONFIGS:
            logger.info("─── Processing: %s ───", cfg.display_name)
            try:
                self._process_scrub_type(cfg, stats,
                                         start_date=start_date,
                                         end_date=end_date)
            except Exception as e:
                msg = f"{cfg.display_name}: unexpected error — {e}"
                logger.error(msg, exc_info=True)
                stats.errors.append(msg)
                stats.files_errored += 1
        return stats

    def run_single(self, scrub_key: str, *,
                   start_date: Optional[datetime.date] = None,
                   end_date: Optional[datetime.date] = None) -> RunStats:
        """Run a single scrub type by key."""
        cfg = get_config(scrub_key)
        stats = RunStats()
        self._process_scrub_type(cfg, stats,
                                 start_date=start_date, end_date=end_date)
        return stats

    def run_from_local_file(self, scrub_key: str,
                            file_path: str,
                            scrub_date: Optional[datetime.date] = None
                            ) -> RunStats:
        """
        Parse a local .xlsx file instead of downloading from SharePoint.
        Useful for testing and manual backfill.
        """
        cfg = get_config(scrub_key)
        stats = RunStats()

        if scrub_date is None:
            scrub_date = extract_date_from_filename(file_path)
        if scrub_date is None:
            stats.errors.append(f"Could not extract date from '{file_path}'")
            return stats

        stats.files_found = 1

        try:
            wb = load_workbook(file_path, data_only=True, read_only=True)
            parser_func = get_parser(cfg.parser_func)
            raw_results = parser_func(wb, file_date=scrub_date)
            wb.close()
        except ParserError as e:
            stats.errors.append(str(e))
            stats.files_errored += 1
            return stats
        except Exception as e:
            stats.errors.append(f"Error parsing '{file_path}': {e}")
            stats.files_errored += 1
            return stats

        stats.files_parsed += 1
        results = self._map_results(cfg, scrub_date, raw_results, file_path)
        self._store_results(results, stats)

        return stats

    def run_from_buffer(self, scrub_key: str,
                        file_buf: io.BytesIO,
                        filename: str,
                        source_url: str = "",
                        scrub_date: Optional[datetime.date] = None
                        ) -> RunStats:
        """
        Parse an in-memory .xlsx buffer (e.g. from a Power Automate POST).
        The scrub_key can be auto-detected upstream — see ingest.detect_scrub_type().
        """
        cfg = get_config(scrub_key)
        stats = RunStats()

        if scrub_date is None:
            scrub_date = extract_date_from_filename(filename)
        if scrub_date is None:
            stats.errors.append(f"Could not extract date from '{filename}'")
            return stats

        stats.files_found = 1

        try:
            wb = load_workbook(file_buf, data_only=True, read_only=True)
            parser_func = get_parser(cfg.parser_func)
            raw_results = parser_func(wb, file_date=scrub_date)
            wb.close()
        except ParserError as e:
            stats.errors.append(str(e))
            stats.files_errored += 1
            return stats
        except Exception as e:
            stats.errors.append(f"Error parsing '{filename}': {e}")
            stats.files_errored += 1
            return stats

        stats.files_parsed += 1
        source = source_url or filename
        results = self._map_results(cfg, scrub_date, raw_results, source)
        self._store_results(results, stats)

        return stats

    # ── Internal pipeline ──────────────────────────────────────────────────

    def _process_scrub_type(self, cfg: ScrubTypeConfig, stats: RunStats, *,
                            start_date: Optional[datetime.date] = None,
                            end_date: Optional[datetime.date] = None):
        """Full pipeline for one scrub type."""

        # 1. Discover files
        files = self._discover_files(cfg)
        stats.files_found += len(files)

        if not files:
            logger.info("No files found for %s", cfg.display_name)
            return

        # 2. Group split files by date (for $0 Product Fee, etc.)
        grouped = self._group_files_by_date(files)

        for scrub_date, file_group in sorted(grouped.items()):
            # date filter
            if start_date and scrub_date < start_date:
                continue
            if end_date and scrub_date > end_date:
                continue

            # 3. Dedup check — skip if ALL outputs already exist for this date
            if self._all_outputs_exist(cfg, scrub_date):
                stats.files_skipped_dedup += len(file_group)
                logger.debug("Skipping %s on %s (already exists)",
                             cfg.display_name, scrub_date)
                continue

            # 4. Download + parse (summing across split files)
            try:
                combined_results = self._download_and_parse(
                    cfg, file_group, scrub_date
                )
            except ParserError as e:
                msg = f"{cfg.display_name} ({scrub_date}): {e}"
                logger.warning(msg)
                stats.errors.append(msg)
                stats.files_errored += len(file_group)
                continue
            except Exception as e:
                msg = f"{cfg.display_name} ({scrub_date}): download/parse error — {e}"
                logger.error(msg, exc_info=True)
                stats.errors.append(msg)
                stats.files_errored += len(file_group)
                continue

            stats.files_parsed += len(file_group)

            # 5. Map to ScrubResult objects
            source_url = file_group[0].web_url
            results = self._map_results(cfg, scrub_date,
                                        combined_results, source_url)

            # 6. Store
            self._store_results(results, stats)

    def _discover_files(self, cfg: ScrubTypeConfig) -> list[SharePointFile]:
        """Find files for a scrub type using folder listing or search."""
        try:
            files = self.graph.list_folder_files(cfg.folder_path)
        except Exception as e:
            logger.warning("Folder listing failed for %s, trying search: %s",
                           cfg.display_name, e)
            files = self.graph.search_files(cfg.search_query)

        # filter by file_pattern if set
        if cfg.file_pattern:
            pattern = re.compile(cfg.file_pattern, re.IGNORECASE)
            files = [f for f in files if pattern.search(f.name)]

        # filter to only files with parseable dates in filename
        valid = []
        for f in files:
            d = extract_date_from_filename(f.name)
            if d is not None:
                valid.append(f)
            else:
                logger.debug("Skipping file (no date in name): %s", f.name)

        return valid

    def _group_files_by_date(self, files: list[SharePointFile]
                             ) -> dict[datetime.date, list[SharePointFile]]:
        """Group files by their extracted date (handles split files)."""
        grouped: dict[datetime.date, list[SharePointFile]] = defaultdict(list)
        for f in files:
            d = extract_date_from_filename(f.name)
            if d:
                grouped[d].append(f)
        return dict(grouped)

    def _all_outputs_exist(self, cfg: ScrubTypeConfig,
                           scrub_date: datetime.date) -> bool:
        """Check if all output categories already have entries for this date."""
        for out in cfg.outputs:
            if not self.db.entry_exists(out.section.value, out.category,
                                        scrub_date):
                return False
        return True

    def _download_and_parse(self, cfg: ScrubTypeConfig,
                            file_group: list[SharePointFile],
                            scrub_date: datetime.date
                            ) -> dict[str, int | Decimal]:
        """
        Download file(s), parse each, and combine results.
        For split files (e.g. $0 Product Fee parts), sums counts across parts.
        """
        parser_func = get_parser(cfg.parser_func)
        combined: dict[str, int | Decimal] = {}

        for sp_file in file_group:
            buf = self.graph.download_file(sp_file)
            wb = load_workbook(buf, data_only=True, read_only=True)

            try:
                raw = parser_func(wb, file_date=scrub_date)
            finally:
                wb.close()
                buf.close()

            # merge: sum numeric values across split files
            for key, val in raw.items():
                if key in combined:
                    existing = combined[key]
                    if isinstance(val, (int, float)) and isinstance(existing, (int, float)):
                        combined[key] = existing + val
                    elif isinstance(val, Decimal) and isinstance(existing, Decimal):
                        combined[key] = existing + val
                    else:
                        combined[key] = val  # last-write-wins for non-numeric
                else:
                    combined[key] = val

        return combined

    def _map_results(self, cfg: ScrubTypeConfig,
                     scrub_date: datetime.date,
                     raw_results: dict[str, int | Decimal],
                     source_url: str) -> list[ScrubResult]:
        """
        Map parser output keys to ScrubResult objects using the config's
        OutputMapping definitions.
        """
        results: list[ScrubResult] = []

        for out in cfg.outputs:
            val = raw_results.get(out.output_key)
            if val is None:
                continue

            if out.is_amount:
                result = ScrubResult(
                    section=out.section,
                    category=out.category,
                    scrub_date=scrub_date,
                    amount=Decimal(str(val)) if not isinstance(val, Decimal) else val,
                    source_file=source_url,
                )
            else:
                result = ScrubResult(
                    section=out.section,
                    category=out.category,
                    scrub_date=scrub_date,
                    value=int(val) if not isinstance(val, int) else val,
                    source_file=source_url,
                )
            results.append(result)

        return results

    def _store_results(self, results: list[ScrubResult], stats: RunStats):
        """Insert results into the database (or log if dry_run)."""
        for r in results:
            if self.dry_run:
                logger.info("[DRY RUN] Would insert: %s / %s / %s → v=%s, a=%s",
                            r.section.value, r.category, r.scrub_date,
                            r.value, r.amount)
                stats.entries_inserted += 1
            else:
                if self.db.insert_entry(r):
                    stats.entries_inserted += 1
                else:
                    stats.entries_skipped += 1
