"""
Tests for NCD Scrub Dashboard parsers.
Creates mock openpyxl workbooks to validate extraction logic.

Run: python -m pytest tests/ -v
  or: python tests/test_parsers.py
"""

import datetime
import sys
from decimal import Decimal
from pathlib import Path

# allow running from project root
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook

from scrub_parser.parsers import (
    parse_account_updater,
    parse_ach_rebill,
    parse_all_hold_reasons,
    parse_billing_date_alignment,
    parse_combined_eligibility,
    parse_daily_dupes,
    parse_fbd,
    parse_incomplete_accounts,
    parse_mass_terms,
    parse_missing_mop,
    parse_nbd,
    parse_no_dependent_products,
    parse_overaged_dependents,
    parse_realm_health,
    parse_srs_aged,
    parse_test_member_accounts,
    parse_zero_product_fee,
    ParserError,
)
from scrub_parser.utils import extract_date_from_filename


# ── Helpers ────────────────────────────────────────────────────────────────

def make_wb_with_summary(rows: list[list], sheet_name="Summary") -> Workbook:
    """Create a workbook with a single sheet populated with the given rows."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r, row_data in enumerate(rows, start=1):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)
    return wb


def make_wb_with_data_sheet(rows: list[list], sheet_name="Data") -> Workbook:
    """Create a workbook with a data sheet (no summary)."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r, row_data in enumerate(rows, start=1):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)
    return wb


# ── Tests ──────────────────────────────────────────────────────────────────

class TestDateExtraction:
    def test_standard_filename(self):
        d = extract_date_from_filename("07.14.2026 Billing Date Alignment Scrub.xlsx")
        assert d == datetime.date(2026, 7, 14)

    def test_date_in_middle(self):
        d = extract_date_from_filename("ACH Rebill Cancel 2 Attempts 06.09.2026.xlsx")
        assert d == datetime.date(2026, 6, 9)

    def test_no_date(self):
        assert extract_date_from_filename("random_file.xlsx") is None

    def test_invalid_date(self):
        assert extract_date_from_filename("13.32.2026 bad date.xlsx") is None


class TestBillingDateAlignment:
    def test_basic(self):
        wb = make_wb_with_summary([
            ["Action", "Count"],
            ["Corrected", 10],
            ["No Action", 15],
            ["Grand Total", 25],
        ])
        result = parse_billing_date_alignment(wb)
        assert result == {"count": 25}


class TestFBD:
    def test_two_sections(self):
        wb = make_wb_with_summary([
            ["FBD in the Past", None],
            ["Action", "Count"],
            ["Corrected", 5],
            ["Grand Total", 5],
            [None, None],
            ["FBD Blank", None],
            ["Action", "Count"],
            ["Fixed", 3],
            ["Grand Total", 3],
        ])
        result = parse_fbd(wb)
        assert result["fbd_past"] == 5
        assert result["fbd_blank"] == 3


class TestNBD:
    def test_two_sections(self):
        wb = make_wb_with_summary([
            ["NBD in the Past", None],
            ["Action", "Count"],
            ["Corrected", 8],
            ["Grand Total", 8],
            [None, None],
            ["NBD Blank", None],
            ["Action", "Count"],
            ["Fixed", 4],
            ["Grand Total", 4],
        ])
        result = parse_nbd(wb)
        assert result["nbd_past"] == 8
        assert result["nbd_blank"] == 4


class TestMassTerms:
    def test_pivot_table(self):
        wb = make_wb_with_summary([
            ["Status", "Count of Member ID", "Sum of Product Amount"],
            ["Active", 500, 15000.00],
            ["Termed", 763, 25662.42],
            ["Grand Total", 1263, 40662.42],
        ])
        result = parse_mass_terms(wb)
        assert result["count"] == 1263
        assert result["amount"] == Decimal("40662.42")


class TestACHRebill:
    def test_total_records(self):
        wb = make_wb_with_summary([
            ["Metric", "Value"],
            ["Total Records", 100],
            ["FPN Now", 100],
        ])
        result = parse_ach_rebill(wb)
        assert result["total_records"] == 100


class TestIncompleteAccounts:
    def test_total_row(self):
        wb = make_wb_with_summary([
            ["Action", "Count"],
            ["Called Agent", 2],
            ["No Action", 2],
            ["Total", 4],
        ])
        result = parse_incomplete_accounts(wb)
        assert result == {"count": 4}


class TestNoDependentProducts:
    def test_total_row(self):
        wb = make_wb_with_summary([
            ["Summary of Actions", None],
            ["Action", "Count"],
            ["Added to dependents", 3],
            ["Total", 3],
        ])
        result = parse_no_dependent_products(wb)
        assert result == {"count": 3}


class TestMissingMOP:
    def test_summary_total(self):
        wb = make_wb_with_summary([
            ["Action", "Count"],
            ["Updated", 2],
            ["Total", 2],
        ])
        result = parse_missing_mop(wb)
        assert result == {"count": 2}

    def test_fallback_data_rows(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"  # not "Summary"
        ws.cell(row=1, column=1, value="Member ID")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=2, column=1, value=1001)
        ws.cell(row=2, column=2, value="John")
        ws.cell(row=3, column=1, value=1002)
        ws.cell(row=3, column=2, value="Jane")
        result = parse_missing_mop(wb)
        assert result == {"count": 2}


class TestSRsAged:
    def test_sum_actions(self):
        wb = make_wb_with_summary([
            ["Action", "Count"],
            ["Followed Up", 10],
            ["Escalated", 8],
            ["Closed", 5],
            ["Pending", 3],
        ])
        result = parse_srs_aged(wb)
        # no Total row → falls back to summing column B
        assert result == {"count": 26}

    def test_with_total(self):
        wb = make_wb_with_summary([
            ["Action", "Count"],
            ["Followed Up", 10],
            ["Escalated", 8],
            ["Total", 18],
        ])
        result = parse_srs_aged(wb)
        assert result == {"count": 18}


class TestTestMemberAccounts:
    def test_basic(self):
        wb = make_wb_with_summary([
            ["Action", "Count"],
            ["Deleted", 1],
            ["Total", 1],
        ])
        result = parse_test_member_accounts(wb)
        assert result == {"count": 1}


class TestOveragedDependents:
    def test_single_section(self):
        wb = make_wb_with_summary([
            ["Action", "Count"],
            ["Termed dependent", 18],
            ["Termed dep rewrite", 4],
            ["Grand Total", 22],
        ])
        result = parse_overaged_dependents(wb)
        assert result["count"] == 22

    def test_split_sections(self):
        wb = make_wb_with_summary([
            ["Final Notice", None],
            ["Action", "Count"],
            ["Termed", 10],
            ["Total", 10],
            [None, None],
            ["30-Day Notice", None],
            ["Action", "Count"],
            ["Sent Notice", 5],
            ["Total", 5],
        ])
        result = parse_overaged_dependents(wb)
        assert result["final_notice"] == 10
        assert result["thirty_day"] == 5
        assert result["count"] == 15


class TestAllHoldReasons:
    def test_filtered_count(self):
        file_date = datetime.date(2026, 7, 14)
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        ws.cell(row=1, column=1, value="Member ID")
        ws.cell(row=1, column=2, value="Hold Date")
        # 3 months ago → should be counted
        ws.cell(row=2, column=1, value=1001)
        ws.cell(row=2, column=2, value=datetime.date(2026, 4, 1))
        # 1 month ago → should NOT be counted
        ws.cell(row=3, column=1, value=1002)
        ws.cell(row=3, column=2, value=datetime.date(2026, 6, 20))
        # 6 months ago → should be counted
        ws.cell(row=4, column=1, value=1003)
        ws.cell(row=4, column=2, value=datetime.date(2026, 1, 10))

        result = parse_all_hold_reasons(wb, file_date=file_date)
        assert result == {"count": 2}


class TestZeroProductFee:
    def test_data_row_count(self):
        wb = make_wb_with_data_sheet([
            ["Member ID", "Product", "Fee"],
            [1001, "Plan A", 0],
            [1002, "Plan B", 0],
            [1003, "Plan C", 0],
        ])
        result = parse_zero_product_fee(wb)
        assert result == {"count": 3}


class TestDailyDupes:
    def test_data_row_count(self):
        wb = make_wb_with_data_sheet([
            ["Member ID", "Name", "Date"],
            [1001, "John", "2026-07-14"],
            [1002, "Jane", "2026-07-14"],
        ])
        result = parse_daily_dupes(wb)
        assert result == {"count": 2}


class TestCombinedEligibility:
    def test_section_boundaries(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        # MetLife section
        ws.cell(row=1, column=1, value="MetLife EDI Errors")
        ws.cell(row=2, column=1, value="Error Type")
        ws.cell(row=2, column=2, value="Split Error Count")
        ws.cell(row=3, column=1, value="Missing SSN")
        ws.cell(row=3, column=2, value=5)
        ws.cell(row=4, column=1, value="Bad DOB")
        ws.cell(row=4, column=2, value=3)
        ws.cell(row=5, column=1, value="Grand Total")
        ws.cell(row=5, column=2, value=8)
        # VSP section
        ws.cell(row=7, column=1, value="VSP EDI Errors")
        ws.cell(row=8, column=1, value="Error Type")
        ws.cell(row=8, column=2, value="Split Error Count")
        ws.cell(row=9, column=1, value="Missing Name")
        ws.cell(row=9, column=2, value=2)
        ws.cell(row=10, column=1, value="Grand Total")
        ws.cell(row=10, column=2, value=2)
        # NWFA section
        ws.cell(row=12, column=1, value="NWFA Errors")
        ws.cell(row=13, column=1, value="Error Type")
        ws.cell(row=13, column=2, value="Split Error Count")
        ws.cell(row=14, column=1, value="Dup Entry")
        ws.cell(row=14, column=2, value=1)
        ws.cell(row=15, column=1, value="Grand Total")
        ws.cell(row=15, column=2, value=1)

        result = parse_combined_eligibility(wb)
        assert result["metlife"] == 8
        assert result["vsp"] == 2
        assert result["nwfa"] == 1


class TestRealmHealth:
    def test_metlife(self):
        wb = make_wb_with_data_sheet([
            ["MetLife EDI Errors", "Name", "Error"],
            [1001, "John", "Missing SSN"],
            [1002, "Jane", "Bad DOB"],
        ])
        result = parse_realm_health(wb)
        assert result == {"metlife_carrier": 2}

    def test_vsp(self):
        wb = make_wb_with_data_sheet([
            ["VSP EDI Errors", "Name", "Error"],
            [1001, "John", "Missing Name"],
        ])
        result = parse_realm_health(wb)
        assert result == {"vsp_carrier": 1}


class TestAccountUpdater:
    def test_billing_summary(self):
        wb = make_wb_with_summary([
            ["Billing Summary", None],
            ["Total Members", 67],
            ["Total Billing", 2084.25],
        ])
        result = parse_account_updater(wb)
        assert result["count"] == 67
        assert result["amount"] == Decimal("2084.25")


# ── Run all tests ──────────────────────────────────────────────────────────

def run_tests():
    """Simple test runner (no pytest required)."""
    import traceback

    test_classes = [
        TestDateExtraction,
        TestBillingDateAlignment,
        TestFBD,
        TestNBD,
        TestMassTerms,
        TestACHRebill,
        TestIncompleteAccounts,
        TestNoDependentProducts,
        TestMissingMOP,
        TestSRsAged,
        TestTestMemberAccounts,
        TestOveragedDependents,
        TestAllHoldReasons,
        TestZeroProductFee,
        TestDailyDupes,
        TestCombinedEligibility,
        TestRealmHealth,
        TestAccountUpdater,
    ]

    passed = 0
    failed = 0
    errors = []

    for cls in test_classes:
        instance = cls()
        methods = [m for m in dir(instance) if m.startswith("test_")]
        for method_name in methods:
            label = f"{cls.__name__}.{method_name}"
            try:
                getattr(instance, method_name)()
                print(f"  ✓ {label}")
                passed += 1
            except Exception as e:
                print(f"  ✗ {label}: {e}")
                errors.append((label, traceback.format_exc()))
                failed += 1

    print(f"\n{'=' * 60}")
    print(f"Results: {passed} passed, {failed} failed")
    if errors:
        print(f"\nFailure details:")
        for label, tb in errors:
            print(f"\n--- {label} ---")
            print(tb)

    return failed == 0


if __name__ == "__main__":
    success = run_tests()
    sys.exit(0 if success else 1)
